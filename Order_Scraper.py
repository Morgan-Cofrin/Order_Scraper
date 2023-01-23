import time
import openpyxl
import calendar
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#Months for summary
months_dict = {"MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER"}

#Filename
file_name = "RangeCreek_Order_Info.xlsx"

#website creds
nhmu_username = "Your_Username_Here"
nhmu_password = "Your_Password_Here"

#Define workbook
workbook = openpyxl.load_workbook(file_name)

# Providing URLs, place holder fstrings
url_login = 'https://nhmu.utah.edu/user'
url_rangecreek_orders = "https://nhmu.utah.edu/rangecreek/orders?page={}"
url_order = "https://nhmu.utah.edu/rangecreek/order/{}"
url_visitor = "https://nhmu.utah.edu/rangecreek/order/visitors/{}?sin=0&onum={}"

# Open a web browser
driver = webdriver.Chrome()

#Class definitions
class Order:
    def __init__(self, order_index, order_number, order_purchaser, date_purchased, order_status, order_quantity, order_total):
        self.order_index = order_index
        self.order_number = order_number
        self.order_purchaser = order_purchaser
        self.date_purchased = date_purchased
        self.order_status = order_status
        self.order_quantity = order_quantity
        self.order_total = order_total

class Visitor:
    def __init__(self, order: Order, first_name, last_name, season_pass_number, date_of_birth, phone_number, address_1, address_2, zip_code, city, state, country, purpose_for_visit):
        self.order_index = order.order_index
        self.order_number = order.order_number
        self.first_name = first_name
        self.last_name = last_name
        self.season_pass = season_pass_number
        self.date_of_birth = date_of_birth
        self.phone_number = phone_number
        self.address_1 = address_1
        self.address_2 = address_2
        self.zip_code = zip_code
        self.city = city
        self.state = state
        self.country = country
        self.purpose_for_visit = purpose_for_visit

class Visit_Dates:
    def __init__(self, visitor: Visitor, visit_date, day_permit_number):
        self.order_index = visitor.order_index
        self.order_number = visitor.order_number
        self.season_pass = visitor.season_pass
        self.visit_date = visit_date
        self.day_permit_number = day_permit_number


#Helpers
def setup_orders_sheet():
    #Check if the worksheet already exists
    try:
        sheet = workbook['ORDERS']
    except KeyError:
        #If it doesn't exist, create a new worksheet named "ORDERS"
        sheet = workbook.create_sheet("ORDERS")

    # check if the headers already exist
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "ORDER_PURCHASER":
        sheet["C1"] = "ORDER_PURCHASER"
    if sheet["D1"].value != "DATE_PURCHASED":
        sheet["D1"] = "DATE_PURCHASED"
    if sheet["E1"].value != "ORDER_STATUS":
        sheet["E1"] = "ORDER_STATUS"
    if sheet["F1"].value != "ORDER_QUANTITY":
        sheet["F1"] = "ORDER_QUANTITY"
    if sheet["G1"].value != "ORDER_TOTAL":
        sheet["G1"] = "ORDER_TOTAL"

    return sheet

def setup_visitors_sheet():
    #Check if the worksheet already exists
    try:
        sheet = workbook['VISITORS']
    except KeyError:
        #If it doesn't exist, create a new worksheet named "VISITORS"
        sheet = workbook.create_sheet("VISITORS")
    
    # check if the headers already exist
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "FIRST_NAME":
        sheet["C1"] = "FIRST_NAME"
    if sheet["D1"].value != "LAST_NAME":
        sheet["D1"] = "LAST_NAME"
    if sheet["E1"].value != "SEASON_PASS_NUMBER":
        sheet["E1"] = "SEASON_PASS_NUMBER"
    if sheet["F1"].value != "DATE_OF_BIRTH":
        sheet["F1"] = "DATE_OF_BIRTH"
    if sheet["G1"].value != "PHONE_NUMBER":
        sheet["G1"] = "PHONE_NUMBER"
    if sheet["H1"].value != "ADDRESS_1":
        sheet["H1"] = "ADDRESS_1"
    if sheet["I1"].value != "ADDRESS_2":
        sheet["I1"] = "ADDRESS_2"
    if sheet["J1"].value != "ZIP_CODE":
        sheet["J1"] = "ZIP_CODE"
    if sheet["K1"].value != "CITY":
        sheet["K1"] = "CITY"
    if sheet["L1"].value != "STATE":
        sheet["L1"] = "STATE"
    if sheet["M1"].value != "COUNTRY":
        sheet["M1"] = "COUNTRY"
    if sheet["N1"].value != "PURPOSE_FOR_VISIT":
        sheet["N1"] = "PURPOSE_FOR_VISIT"

    return sheet

def setup_visit_dates_sheet():

    #Check if the worksheet already exists
    try:
        sheet = workbook['VISIT_DATES']
    except KeyError:
        #If it doesn't exist, create a new worksheet named 'VISIT_DATES'
        sheet = workbook.create_sheet('VISIT_DATES')

    #Check if headers exist, and add them if not
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "SEASON_PASS":
        sheet["C1"] = "SEASON_PASS"
    if sheet["D1"].value != "VISIT_DATE":
        sheet["D1"] = "VISIT_DATE"
    if sheet["E1"].value != "DAY_PERMIT_NUMBER":
        sheet["E1"] = "DAY_PERMIT_NUMBER"

    return sheet

def setup_summary_sheet():
    # Check if sheet already exists in workbook
    if "PASS_PERMIT_SUMMARY" in workbook.sheetnames:
        sheet = workbook["PASS_PERMIT_SUMMARY"]
    else:
        sheet = workbook.create_sheet("PASS_PERMIT_SUMMARY")
    
    # Check if headers already exist
    if sheet["A1"].value != "YEAR" or sheet["B1"].value != "NO. PASSES" or sheet["D1"].value != "YEAR" \
    or sheet["E1"].value != "MONTH" or sheet["F1"].value != "NO. PERMITS":

        sheet.merge_cells("A1:B1")
        sheet["A1"] = "SEASON PASSES"
        sheet.merge_cells("D1:F1")
        sheet["D1"] = "DAY PERMITS (non-commercial)"
        sheet["A2"] = "YEAR"
        sheet["B2"] = "NO. PASSES"
        sheet["D2"] = "YEAR"
        sheet["E2"] = "MONTH"
        sheet["F2"] = "NO. PERMITS"

    return sheet


#Parsing Info
def read_order_info(order_row):

    next_order = find_next_order(order_row)

    #Navigate to order
    driver.get(next_order[0])

    # find the elements by ID
    divs = driver.find_elements(By.CLASS_NAME, 'col-md-12')
    text = divs[1].text

    order_index = next_order[1]
    order_number = next_order[2]
    order_purchaser = text.split("Purchaser: ")[1].split("\n")[0]
    date_purchased = text.split("Date: ")[1].split("\n")[0]
    order_status = text.split("Status: ")[1].split("\n")[0]
    order_quantity = text.split("Quantity: ")[1].split("\n")[0]
    order_total = text.split("Total: ")[1].split("\n")[0]

    return Order(
        order_index,
        order_number,
        order_purchaser,
        date_purchased,
        order_status,
        order_quantity,
        order_total,
    )

def read_visitor_info(order):
    #Navigate to Visitor page
    driver.get((url_visitor).format(order.order_index, order.order_number))

    #Find ALL elements by ID
    panel_bodies = driver.find_elements(By.CLASS_NAME, 'panel-body')
    visitors = []

    for panel_body in panel_bodies:
        #Find the element with class 'col-md-4' within the current panel_body
        col_md_4 = panel_body.find_element(By.CLASS_NAME, "col-md-4")
        text = col_md_4.text
        first_name = text.split("First Name:")[1].split("\n")[0]
        last_name = text.split("Last Name:")[1].split("\n")[0]
        season_pass_number = text.split("Season Pass:")[1].split("\n")[0]
        date_of_birth = text.split("DOB:")[1].split("\n")[0]
        phone_number = text.split("Phone Number:")[1].split("\n")[0]
        address_1 = text.split("Address 1:")[1].split("\n")[0]
        address_2 = text.split("Address 2:")[1].split("\n")[0]
        zip_code = text.split("Zip Code:")[1].split("\n")[0]
        city = text.split("City:")[1].split("\n")[0]
        state = text.split("State:")[1].split("\n")[0]
        country = text.split("Country:")[1].split("\n")[0]

        #Weird becusae this lives somewhere slightly different on the page. The only <p> tag in <div class=panel-body>
        purpose_for_visit = panel_body.find_element(By.TAG_NAME, "p").text

        #Create a Visitor object
        visitor = Visitor(order, first_name, last_name, season_pass_number, date_of_birth, phone_number, address_1, address_2, zip_code, city, state, country, purpose_for_visit)
        visitors.append(visitor)

    return visitors

def read_visit_date_info(visitors_list):
    
    #Find ALL  table elements by ID
    tables = driver.find_elements(By.CLASS_NAME, 'table')
    visit_dates = []
    i = 0

    for table in tables:
        # check if there is a <tbody> tag
        try:
            tbody = table.find_element(By.TAG_NAME, "tbody")
        except Exception:
            tbody = table
            no_visits_message = tbody.find_element(By.XPATH, "//td[text()='No visits have been scheduled.']")
            visit_dates = Visit_Dates(visitors_list[i], date=no_visits_message, day_permit_number="")
            visit_dates.append(visit_dates)
            i +=1

        #Find all rows in the table
        rows = tbody.find_elements(By.TAG_NAME, 'tr')
        for row in rows:
            #Find all cells in the row
            cells = row.find_elements(By.TAG_NAME, 'td')
            if len(cells)==2:
                date = cells[0].text
                day_permit_number = cells[1].text
                visit_date = Visit_Dates(visitors_list[i], date, day_permit_number)
                visit_dates.append(visit_date)
        i +=1

    return visit_dates

def parse_dates_summary(sheet, column_index):
    # sourcery skip: for-append-to-extend, inline-immediately-returned-variable, list-comprehension, use-contextlib-suppress
    year_dict = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        date = row[column_index]
        try:
            if isinstance(date, str):
                date = datetime.strptime(date, '%m-%d-%Y')
            month, year = date.month, date.year
            if year not in year_dict:
                year_dict[year] = {month: 0 for month in range(5, 12)}
            if 5 <= month <= 11:
                year_dict[year][month] += 1
        except ValueError:
            pass

    sorted_years = sorted(year_dict.keys(), reverse=True)
    sorted_data = []
    for year in sorted_years:
        sorted_data.append({year: year_dict[year]})

    return sorted_data


#Writing info to Excel
def write_order_to_excel(order):

    #Check if sheet/headers exist, and add them if not
    orders_sheet = setup_orders_sheet()

    #Check if the order already exists in the worksheet
    for row in orders_sheet.iter_rows(values_only=True):
        if row[0] == order.order_index and row[1] == order.order_number:
            print(f'Order {order.order_number} already exists in the worksheet.')
            break
    else:
        #If the order doesn't exist, add it to the worksheet
        orders_sheet.append([order.order_index, order.order_number, order.order_purchaser, order.date_purchased, order.order_status, order.order_quantity, order.order_total])
        #workbook.save(file_name) #saving after order + visitor + dates instead, keeps relations intact
        print(f'Order {order.order_number} added to the worksheet')

def write_visitor_to_excel(visitors_list):

    #Check if sheet/headers exist, and add them if not
    visitors_sheet = setup_visitors_sheet()

    for visitor in visitors_list:
        #Check if visitor already exists
        for row in visitors_sheet.iter_rows(values_only=True):
            if row[0] == visitor.order_index and row[1] == visitor.order_number and row[2] == visitor.first_name and row[3] == visitor.last_name:  #TODO could potentially reference season pass number instead
                print(f'Visitor {visitor.first_name} {visitor.last_name} already exists in the worksheet.')
                break
        else:
            #If the visitor doesn't exist, add it to the worksheet
            visitors_sheet.append([visitor.order_index, visitor.order_number, visitor.first_name, visitor.last_name, visitor.season_pass, visitor.date_of_birth, 
            visitor.phone_number, visitor.address_1, visitor.address_2, visitor.zip_code, visitor.city, visitor.state, visitor.country, visitor.purpose_for_visit])
            #workbook.save(file_name) #saving after order + visitor + dates instead, keeps relations intact
            print(f'Visitor {visitor.first_name} {visitor.last_name} added to the worksheet.')

def write_visit_date_to_excel(visit_dates_list):

    #Check if sheet/headers exist, and add them if not
    visit_dates_sheet = setup_visit_dates_sheet()

    for visit_date in visit_dates_list:
        #Check if visit date already exists
        for row in visit_dates_sheet.iter_rows(values_only=True):
            if row[0] == visit_date.order_index and row[1] == visit_date.order_number and row[2] == visit_date.season_pass and row[4] == visit_date.day_permit_number:
                print(f'Visit date "{visit_date.season_pass}, {visit_date.day_permit_number}" already exists in the worksheet.')
                break
        else:
            #If the visitor doesn't exist, add it to the worksheet
            visit_dates_sheet.append([visit_date.order_index, visit_date.order_number, visit_date.season_pass, visit_date.visit_date, visit_date.day_permit_number])
            #workbook.save(file_name) #saving after order + visitor + dates instead, keeps relations intact
            print(f'Visit date "{visit_date.season_pass}, {visit_date.day_permit_number}" added to the worksheet.')

def write_summary_to_excel():
    # Clear sheet, create fresh one. Can't figure out how to update existing values
    # Also, try: Find/Select, Goto: special, delete blank cells in column A + B, shift cells up
    try: 
        workbook.remove('PASS_PERMIT_SUMMARY')
        summary_sheet = setup_summary_sheet()
    except Exception:
        summary_sheet = setup_summary_sheet()

    sheet_to_parse = workbook['VISIT_DATES']
    column_to_parse = 3

    data = parse_dates_summary(sheet_to_parse, column_to_parse)

    for i, year in enumerate(data):
        year_total = 0
        for year_key in year:
            for month_key in year[year_key]:
                if month_key in months_dict:
                    month_name = month_key
                else:
                    month_name = calendar.month_name[month_key].upper()
                summary_sheet.cell(row=summary_sheet.max_row + 1, column=4, value=year_key)
                summary_sheet.cell(row=summary_sheet.max_row, column=5, value=month_name)
                summary_sheet.cell(row=summary_sheet.max_row, column=6, value=year[year_key][month_key])
                year_total += year[year_key][month_key] # adding the number of permit for that month to the total for that year
        summary_sheet.cell(row=summary_sheet.max_row + 1, column=1, value=year_key)
        summary_sheet.cell(row=summary_sheet.max_row, column=2, value=year_total)
        if i < len(data)-1:
            summary_sheet.cell(row=summary_sheet.max_row + 1, column=4, value=None)

    workbook.save(file_name)


#Main stuff
def get_everything_ready():    
    # Navigate to the page
    driver.get(url_login)

    # Wait for the username and password fields to be present on the page
    wait = WebDriverWait(driver, 3)

    username_field = wait.until(EC.presence_of_element_located((By.ID, "edit-name")))
    password_field = wait.until(EC.presence_of_element_located((By.ID, "edit-pass")))

    # Enter the username and password
    username_field.send_keys(nhmu_username)
    time.sleep(2)
    password_field.send_keys(nhmu_password)
    time.sleep(2)

    # Find the form submit button and click it
    submit_button = driver.find_element(By.ID, "edit-submit")
    submit_button.click()
    time.sleep(1)

def collect_all_orders_on_page():

    #maybe can change this to get all href from the visitors button and parse that instead of row objects
    all_order_rows = driver.find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
    list_of_order_hrefs = []
    for row in all_order_rows:
        href = row.find_element(By.TAG_NAME, 'a').get_attribute('href')
        list_of_order_hrefs.append(href)


    return list_of_order_hrefs

def all_reads_and_writes(order_row):
    ##ALL WORKS, DON'T TOUCH 
    # mabye change the order of the things but probs just picky, or might die
    order_obj = read_order_info(order_row)
    write_order_to_excel(order_obj)

    visitors_list = read_visitor_info(order_obj)
    write_visitor_to_excel(visitors_list)

    visit_dates_list = read_visit_date_info(visitors_list)
    write_visit_date_to_excel(visit_dates_list)

    workbook.save(file_name)
    print(f'Saved {file_name}')

def find_next_order(order_row):

    order_index = order_row.rsplit('/', 1)[1].rsplit('?', 1)[0]
    order_number = order_row.rsplit('onum=', 1)[1]
    next_order_url = url_order.format(order_index)

    return next_order_url, order_index, order_number



# -- MAIN -- #

#Access the Range Creek Page with nhmu.utah.edu/user
get_everything_ready()


more_orders = True
i = 1
# Navigate to Orders, Page 1
driver.get(url_rangecreek_orders.format(i))

while more_orders:
    try:
        if driver.find_element(By.CLASS_NAME, 'table-responsive'):
            #Read all Order URLs
            list_of_order_hrefs = collect_all_orders_on_page()
            print(f"Starting page: {i}")
            #Eat all the Order data and write to Excel
            for order_row in list_of_order_hrefs:
                all_reads_and_writes(order_row)
                print(f"Finished row: {(list_of_order_hrefs.index(order_row))+1} Page: {i}")
            #Goto Next page of orders
            i += 1
            driver.get(url_rangecreek_orders.format((i)))
    except Exception:
        more_orders = False


#Refresh workbook and write summary
workbook.close()
workbook = openpyxl.load_workbook(file_name)
write_summary_to_excel()


#Clean up
driver.close()
workbook.save(file_name)
workbook.close()
