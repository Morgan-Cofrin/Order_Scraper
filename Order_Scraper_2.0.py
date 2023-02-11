import time
import os
import openpyxl
import calendar
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import tkinter as tk
from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox

# Last Updated
today = datetime.now().strftime("%Y-%m-%d")

# Months dictionary for summary
months_dict = {"MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER"}

# Filename
file_name = "RangeCreek_Order_Info.xlsx"

#Define workbook
if not os.path.exists(file_name):
    # Create a new workbook if it doesn't exist
    workbook = openpyxl.Workbook()
    workbook.save(file_name)
    print(f"{file_name} created.")
else:
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_name)
    print(f"{file_name} loaded.")

# Providing URLs, place holder fstrings
url_login = "https://nhmu.utah.edu/user"
url_rangecreek_orders = "https://nhmu.utah.edu/rangecreek/orders?page={}"
url_order = "https://nhmu.utah.edu/rangecreek/order/{}"
url_visitor = "https://nhmu.utah.edu/rangecreek/order/visitors/{}?sin=0&onum={}"


# Class definitions
class Order:
    def __init__(self, order_index, order_number, order_purchaser, date_purchased, order_status, order_quantity, order_total):

        self.order_index = order_index
        self.order_number = order_number
        self.order_purchaser = order_purchaser
        self.date_purchased = date_purchased
        self.order_status = order_status
        self.order_quantity = order_quantity
        self.order_total = order_total

class Visitor:
    def __init__(self, order: Order, first_name, last_name, season_pass_number, date_of_birth, 
                phone_number, address_1, address_2, zip_code, city, state, country, purpose_for_visit):

        self.order_index = order.order_index
        self.order_number = order.order_number
        self.first_name = first_name
        self.last_name = last_name
        self.season_pass_number = season_pass_number
        self.date_of_birth = date_of_birth
        self.phone_number = phone_number
        self.address_1 = address_1
        self.address_2 = address_2
        self.zip_code = zip_code
        self.city = city
        self.state = state
        self.country = country
        self.purpose_for_visit = purpose_for_visit

class Visit_Dates:
    def __init__(self, visitor: Visitor, visit_date, day_permit_number):

        self.order_index = visitor.order_index
        self.order_number = visitor.order_number
        self.season_pass_number = visitor.season_pass_number
        self.visit_date = visit_date
        self.day_permit_number = day_permit_number


# Helpers
def setup_orders_sheet():
    # Check if the worksheet already exists
    try:
        sheet = workbook['ORDERS']
    except KeyError:
        # If it doesn't exist, create a new worksheet named "ORDERS"
        sheet = workbook.create_sheet("ORDERS")

    # Check if the headers already exist
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "ORDER_PURCHASER":
        sheet["C1"] = "ORDER_PURCHASER"
    if sheet["D1"].value != "DATE_PURCHASED":
        sheet["D1"] = "DATE_PURCHASED"
    if sheet["E1"].value != "ORDER_STATUS":
        sheet["E1"] = "ORDER_STATUS"
    if sheet["F1"].value != "ORDER_QUANTITY":
        sheet["F1"] = "ORDER_QUANTITY"
    if sheet["G1"].value != "ORDER_TOTAL":
        sheet["G1"] = "ORDER_TOTAL"
    if sheet["I1"].value != "LAST_UPDATED":
        sheet["I1"] = "LAST_UPDATED"

    return sheet

def setup_visitors_sheet():
    # Check if the worksheet already exists
    try:
        sheet = workbook['VISITORS']
    except KeyError:
        # If it doesn't exist, create a new worksheet named "VISITORS"
        sheet = workbook.create_sheet("VISITORS")
    
    # Check if the headers already exist
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "FIRST_NAME":
        sheet["C1"] = "FIRST_NAME"
    if sheet["D1"].value != "LAST_NAME":
        sheet["D1"] = "LAST_NAME"
    if sheet["E1"].value != "SEASON_PASS_NUMBER":
        sheet["E1"] = "SEASON_PASS_NUMBER"
    if sheet["F1"].value != "DATE_OF_BIRTH":
        sheet["F1"] = "DATE_OF_BIRTH"
    if sheet["G1"].value != "PHONE_NUMBER":
        sheet["G1"] = "PHONE_NUMBER"
    if sheet["H1"].value != "ADDRESS_1":
        sheet["H1"] = "ADDRESS_1"
    if sheet["I1"].value != "ADDRESS_2":
        sheet["I1"] = "ADDRESS_2"
    if sheet["J1"].value != "ZIP_CODE":
        sheet["J1"] = "ZIP_CODE"
    if sheet["K1"].value != "CITY":
        sheet["K1"] = "CITY"
    if sheet["L1"].value != "STATE":
        sheet["L1"] = "STATE"
    if sheet["M1"].value != "COUNTRY":
        sheet["M1"] = "COUNTRY"
    if sheet["N1"].value != "PURPOSE_FOR_VISIT":
        sheet["N1"] = "PURPOSE_FOR_VISIT"
    if sheet["P1"].value != "LAST_UPDATED":
        sheet["P1"] = "LAST_UPDATED"

    return sheet

def setup_visit_dates_sheet():

    # Check if the worksheet already exists
    try:
        sheet = workbook['VISIT_DATES']
    except KeyError:
        # If it doesn't exist, create a new worksheet named 'VISIT_DATES'
        sheet = workbook.create_sheet('VISIT_DATES')

    # Check if headers exist, and add them if not
    if sheet["A1"].value != "ORDER_INDEX":
        sheet["A1"] = "ORDER_INDEX"
    if sheet["B1"].value != "ORDER_NUMBER":
        sheet["B1"] = "ORDER_NUMBER"
    if sheet["C1"].value != "SEASON_PASS_NUMBER":
        sheet["C1"] = "SEASON_PASS_NUMBER"
    if sheet["D1"].value != "VISIT_DATE":
        sheet["D1"] = "VISIT_DATE"
    if sheet["E1"].value != "DAY_PERMIT_NUMBER":
        sheet["E1"] = "DAY_PERMIT_NUMBER"
    if sheet["G1"].value != "LAST_UPDATED":
        sheet["G1"] = "LAST_UPDATED"

    return sheet

def setup_summary_sheet():
    # Check if sheet already exists in workbook
    if "PASS_PERMIT_SUMMARY" in workbook.sheetnames:
        sheet = workbook["PASS_PERMIT_SUMMARY"]
    else:
        sheet = workbook.create_sheet("PASS_PERMIT_SUMMARY")
    
    # Check if headers already exist
    if sheet["A1"].value != "YEAR" or sheet["B1"].value != "NO. PASSES" or sheet["D1"].value != "YEAR" \
    or sheet["E1"].value != "MONTH" or sheet["F1"].value != "NO. PERMITS" or sheet["H1"].value != "LAST_UPDATED" :

        sheet.merge_cells("A1:B1")
        sheet["A1"] = "SEASON PASSES"
        sheet.merge_cells("D1:F1")
        sheet["D1"] = "DAY PERMITS (non-commercial)"
        sheet["A2"] = "YEAR"
        sheet["B2"] = "NO. PASSES"
        sheet["D2"] = "YEAR"
        sheet["E2"] = "MONTH"
        sheet["F2"] = "NO. PERMITS"
        sheet["H1"] = "LAST_UPDATED"

    return sheet


# Parsing Info
def read_order_info(order_tuple):

    # Navigate to the Order via order_index
    driver.get(url_order.format(order_tuple[0]))

    # Find the elements by ID, returns 2 WebElements
    divs = driver.find_elements(By.CLASS_NAME, 'col-md-12')
    # We only want the second instance of 'col-md-12'
    text = divs[1].text

    order_index = order_tuple[0]
    order_number = order_tuple[1]
    order_purchaser = text.split("Purchaser: ")[1].split("\n")[0]
    date_purchased = text.split("Date: ")[1].split("\n")[0]
    order_status = text.split("Status: ")[1].split("\n")[0]
    order_quantity = text.split("Quantity: ")[1].split("\n")[0]
    order_total = text.split("Total: ")[1].split("\n")[0]

    return Order(order_index, order_number, order_purchaser, date_purchased, order_status, order_quantity, order_total)

def read_visitor_info(order):
    # Navigate to Visitor page
    driver.get((url_visitor).format(order.order_index, order.order_number))

    # Find ALL elements by ID
    panel_bodies = driver.find_elements(By.CLASS_NAME, 'panel-body')
    visitors_list = []

    for panel_body in panel_bodies:
        # Find the element with class 'col-md-4' within the current panel_body
        col_md_4 = panel_body.find_element(By.CLASS_NAME, "col-md-4")
        text = col_md_4.text
        first_name = text.split("First Name:")[1].split("\n")[0]
        last_name = text.split("Last Name:")[1].split("\n")[0]
        season_pass_number = text.split("Season Pass:")[1].split("\n")[0]
        date_of_birth = text.split("DOB:")[1].split("\n")[0]
        phone_number = text.split("Phone Number:")[1].split("\n")[0]
        address_1 = text.split("Address 1:")[1].split("\n")[0]
        address_2 = text.split("Address 2:")[1].split("\n")[0]
        zip_code = text.split("Zip Code:")[1].split("\n")[0]
        city = text.split("City:")[1].split("\n")[0]
        state = text.split("State:")[1].split("\n")[0]
        country = text.split("Country:")[1].split("\n")[0]

        # Weird because this lives somewhere slightly different on the page. The only <p> tag in <div class=panel-body>
        purpose_for_visit = panel_body.find_element(By.TAG_NAME, "p").text

        # Create a Visitor object
        visitor = Visitor(order, first_name, last_name, season_pass_number, date_of_birth, phone_number, 
                            address_1, address_2, zip_code, city, state, country, purpose_for_visit)
        visitors_list.append(visitor)

    return visitors_list

def read_visit_date_info(visitors_list):
    
    # Find ALL table elements by ID
    tables = driver.find_elements(By.CLASS_NAME, 'table')
    visit_dates = []
    i = 0

    for table in tables:
        # check if there is a <tbody> tag
        try:
            tbody = table.find_element(By.TAG_NAME, "tbody")
        except Exception:
            tbody = table
            no_visits_message = tbody.find_element(By.XPATH, "//td[text()='No visits have been scheduled.']")
            visit_dates = Visit_Dates(visitors_list[i], date=no_visits_message, day_permit_number="")
            visit_dates.append(visit_dates)
            i +=1

        #Find all rows in the table
        rows = tbody.find_elements(By.TAG_NAME, 'tr')
        for row in rows:
            #Find all cells in the row
            cells = row.find_elements(By.TAG_NAME, 'td')
            if len(cells)==2:
                date = cells[0].text
                day_permit_number = cells[1].text
                visit_date = Visit_Dates(visitors_list[i], date, day_permit_number)
                visit_dates.append(visit_date)
        i +=1

    return visit_dates

def parse_dates_summary(sheet, column_index):
    # sourcery skip: for-append-to-extend, inline-immediately-returned-variable, list-comprehension, use-contextlib-suppress
    year_dict = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        date = row[column_index]
        try:
            if isinstance(date, str):
                date = datetime.strptime(date, '%m-%d-%Y')
            month, year = date.month, date.year
            if year not in year_dict:
                year_dict[year] = {month: 0 for month in range(5, 12)}
            if 5 <= month <= 11:
                year_dict[year][month] += 1
        except ValueError:
            pass

    sorted_years = sorted(year_dict.keys(), reverse=True)
    sorted_data = []
    for year in sorted_years:
        sorted_data.append({year: year_dict[year]})

    return sorted_data


# Writing info to Excel
def write_order_to_excel(order):

    #Check if sheet/headers exist, and add them if not
    orders_sheet = setup_orders_sheet()

    #Check if the order already exists
    for row in orders_sheet.iter_rows(values_only=True):
        if row[0] == order.order_index and row[1] == order.order_number:
            print(f'Order {order.order_number} already exists in the worksheet.')
            break
    else:
        #If the order doesn't exist, add it to the worksheet. Keep newest orders on top, insert row beneath the header
        orders_sheet.insert_rows(2)
        
        # Write to Excel
        try: # Collapsable
            orders_sheet['A2'] = order.order_index
            orders_sheet['B2'] = order.order_number
            orders_sheet['C2'] = order.order_purchaser
            orders_sheet['D2'] = order.date_purchased
            orders_sheet['E2'] = order.order_status
            orders_sheet['F2'] = order.order_quantity
            orders_sheet['G2'] = order.order_total
        except Exception:
            print(f'Couldn\'t write {order.order_number} to sheet')

        print(f'Order {order.order_number} added to the worksheet')

def write_visitor_to_excel(visitors_list):

    #Check if sheet/headers exist, and add them if not
    visitors_sheet = setup_visitors_sheet()

    for visitor in visitors_list:
        #Check if visitor already exists
        for row in visitors_sheet.iter_rows(values_only=True):
            if row[0] == visitor.order_index and row[1] == visitor.order_number and row[4] == visitor.season_pass_number: 
                print(f'Visitor {visitor.first_name} {visitor.last_name} already exists in the worksheet.')
                break
        else:
            #If the visitor doesn't exist, add it to the worksheet
            visitors_sheet.insert_rows(2)

            # Write to Excel
            try: # Collapsable
                visitors_sheet['A2'] = visitor.order_index
                visitors_sheet['B2'] = visitor.order_number
                visitors_sheet['C2'] = visitor.first_name
                visitors_sheet['D2'] = visitor.last_name
                visitors_sheet['E2'] = visitor.season_pass_number
                visitors_sheet['F2'] = visitor.date_of_birth
                visitors_sheet['G2'] = visitor.phone_number
                visitors_sheet['H2'] = visitor.address_1
                visitors_sheet['I2'] = visitor.address_2
                visitors_sheet['J2'] = visitor.zip_code
                visitors_sheet['K2'] = visitor.city
                visitors_sheet['L2'] = visitor.state
                visitors_sheet['M2'] = visitor.country
                visitors_sheet['N2'] = visitor.purpose_for_visit
            except Exception:
                print(f'Couldn\'t write {visitor.first_name} {visitor.last_name} to sheet')

            print(f'Visitor {visitor.first_name} {visitor.last_name} added to the worksheet.')

def write_visit_date_to_excel(visit_dates_list):

    #Check if sheet/headers exist, and add them if not
    visit_dates_sheet = setup_visit_dates_sheet()

    for visit_date in visit_dates_list:
        #Check if visit date already exists
        for row in visit_dates_sheet.iter_rows(values_only=True):
            if row[0] == visit_date.order_index and row[1] == visit_date.order_number and row[2] == visit_date.season_pass_number and row[4] == visit_date.day_permit_number:
                print(f'Visit date "{visit_date.season_pass_number}, {visit_date.day_permit_number}" already exists in the worksheet.')
                break
        else:
            #If the visitor date doesn't exist, add it to the worksheet
            visit_dates_sheet.insert_rows(2)

            # Write to Excel
            try: # Collapsable
                visit_dates_sheet['A2'] = visit_date.order_index
                visit_dates_sheet['B2'] = visit_date.order_number
                visit_dates_sheet['C2'] = visit_date.season_pass
                visit_dates_sheet['D2'] = visit_date.visit_date
                visit_dates_sheet['E2'] = visit_date.day_permit_number
            except Exception:
                print(f'Couldn\'t write {visit_date.season_pass_number} {visit_date.day_permit_number} to the worksheet.')
            
            print(f'Visit date {visit_date.season_pass_number} {visit_date.day_permit_number} added to the worksheet.')

def write_summary_to_excel():
    print("Writing Summary...")

    # Clear sheet, create fresh one. Can't figure out how to update existing values
    # Also, try: Find/Select, Goto: special, delete blank cells in column A + B, shift cells up
    try: 
        workbook.remove('PASS_PERMIT_SUMMARY')
        summary_sheet = setup_summary_sheet()
    except Exception:
        summary_sheet = setup_summary_sheet()

    sheet_to_parse = workbook['VISIT_DATES']
    column_to_parse = 3

    data = parse_dates_summary(sheet_to_parse, column_to_parse)

    for i, year in enumerate(data):
        year_total = 0
        for year_key in year:
            for month_key in year[year_key]:
                if month_key in months_dict:
                    month_name = month_key
                else:
                    month_name = calendar.month_name[month_key].upper()
                summary_sheet.cell(row=summary_sheet.max_row + 1, column=4, value=year_key)
                summary_sheet.cell(row=summary_sheet.max_row, column=5, value=month_name)
                summary_sheet.cell(row=summary_sheet.max_row, column=6, value=year[year_key][month_key])
                year_total += year[year_key][month_key] # adding the number of permit for that month to the total for that year
        summary_sheet.cell(row=summary_sheet.max_row + 1, column=1, value=year_key)
        summary_sheet.cell(row=summary_sheet.max_row, column=2, value=year_total)
        if i < len(data)-1:
            summary_sheet.cell(row=summary_sheet.max_row + 1, column=4, value=None)

    workbook.save(file_name)


# Main stuff
def get_webpage_ready():

    # Navigate to the page
    driver.get(url_login)

    # Wait for the username and password fields to be present on the page
    wait = WebDriverWait(driver, 3)

    username_field = wait.until(EC.presence_of_element_located((By.ID, "edit-name")))
    password_field = wait.until(EC.presence_of_element_located((By.ID, "edit-pass")))

    # Enter the username and password, have to sleep or the site thinks we are running a script or something
    username_field.send_keys(nhmu_username)
    time.sleep(1)
    password_field.send_keys(nhmu_password)
    time.sleep(1)

    # Find the form submit button and click it
    submit_button = driver.find_element(By.ID, "edit-submit")
    submit_button.click()
    time.sleep(1)

    # If creds are bad, close the driver and reopen creds GUI and try again
    if driver.current_url == url_login:
        tkinter_gui()
        get_webpage_ready()
        
def collect_all_orders():

    list_of_all_orders = []
    more_orders = True
    i = 1
    
    sheet = workbook['ORDERS']
    order_index_colulmn = sheet['A']
    order_number_column = sheet['B']

    # Navigate to Orders, Page 1
    driver.get(url_rangecreek_orders.format(i))
    while more_orders:
        try:
            if driver.find_element(By.CLASS_NAME, 'table-responsive'):

                #Read all Order URLs
                all_order_rows = driver.find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
                
                # Parse href from "Visitors" button link into order_number and order_index
                for row in all_order_rows:
                    href = row.find_element(By.TAG_NAME, 'a').get_attribute('href')
                    order_index = href.rsplit('/', 1)[1].rsplit('?', 1)[0]
                    order_number = href.rsplit('onum=', 1)[1]
                    
                    # Check if the order_index and order_number are already in the sheet
                    order_index_exists = False
                    order_number_exists = False
                    for cell in order_index_colulmn:
                        if cell.value == order_index:
                            order_index_exists = True
                            break
                    for cell in order_number_column:
                        if cell.value == order_number:
                            order_number_exists = True
                            break

                    # Append the values to the list_of_all_orders only if they are not already in the sheet
                    if not order_index_exists and not order_number_exists:
                        list_of_all_orders.append((order_index, order_number))
                    else:
                        # Stop checking for more orders when you find one already in the sheet
                        # All order numbers in the sheet and on the site are always highest to lowest, if you find one, the rest will be there too
                        more_orders = False
                        break

                #Goto Next page of orders
                i += 1
                driver.get(url_rangecreek_orders.format(i))
        
        except NoSuchElementException:
            # No table = No more orders
            more_orders = False

    # FILO - becuase there is no sort function for Openpyxl
    # "Oldest" of the new Orders goes in the spreadsheet row 2, and so on and so on, so they stay in order (no pun intended)
    list_of_all_orders.reverse()
    return list_of_all_orders

def all_reads_and_writes(order_tuple):
    ##ALL WORKS, DON'T TOUCH 
    order_obj = read_order_info(order_tuple)
    write_order_to_excel(order_obj)

    visitors_list = read_visitor_info(order_obj)
    write_visitor_to_excel(visitors_list)

    visit_dates_list = read_visit_date_info(visitors_list)
    write_visit_date_to_excel(visit_dates_list)

    workbook.save(file_name)
    print(f'Saved {file_name}')

def tkinter_gui():
    root = tk.Tk()

    root.title("Range Creek Order Scraper")
    root.geometry("400x240")
    # Set background color
    root.config(bg="#bd4f19")

    # Website creds
    input_username = StringVar()
    input_password = StringVar()

    def set_username_password():
        # Website Creds, global so they can be used outside of Tkinter
        global nhmu_username
        global nhmu_password
        nhmu_username = input_username.get()
        nhmu_password = input_password.get()
        messagebox.showinfo(title="Order Scraper Automation", 
        message=("Automation software will begin running now.\n" + 
                "Estimated wait time: 10 mins per 50 new orders.\n\n" +
                "You can continue to use your machine, just don\'t close the automated browser or open the RangeCreek_Order_Info workbook"))
        root.destroy()

    def on_closing(): 
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            root.destroy()
            quit()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    label_usename = Label(root, text="NHMU website username:").pack()
    entry_username = Entry(root, justify="left", textvariable=input_username).pack()

    label_password = Label(root, text="NHMU website password:").pack()
    entry_password = Entry(root, justify="left", textvariable=input_password, show="*").pack()

    btn_submit = Button(root,  text="Submit", command=set_username_password).pack()

    root.mainloop()


# -- MAIN -- #


# -- GUI portion to get creds from user -- #
tkinter_gui()


# -- Selenium portion to do all the things once inside the webpage -- #
# Open the web browser
driver = webdriver.Chrome()

#Access the Range Creek Page with nhmu.utah.edu/user
get_webpage_ready()

# Find all Orders, return list of tuples (order_index, order_number) from href in row.
list_of_new_orders = collect_all_orders()

try:
    #Eat all the Order data and write to Excel
    for order_tuple in list_of_new_orders:
        all_reads_and_writes(order_tuple)
        print(f"Finished row: {(list_of_new_orders.index(order_tuple))+1}")

    #Refresh workbook and write summary
    workbook.close()
    workbook = openpyxl.load_workbook(file_name)
    write_summary_to_excel()
except IndexError:
    print("There are no new orders to read.")


#Clean up
driver.close()
workbook.save(file_name)
print('Saving {}...'.format(file_name))
workbook.close()
